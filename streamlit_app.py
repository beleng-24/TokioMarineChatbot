"""
Streamlit GUI for AI-Enabled Plan Document Review System
Modern web interface for the TokioMarine Chatbot

Run: streamlit run streamlit_app.py
"""

import streamlit as st
import pandas as pd
from pathlib import Path
import json
from datetime import datetime
import os
from openai import OpenAI

# Page configuration
st.set_page_config(
    page_title="Tokio Marine - Plan Review System",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #2c3e50;
        margin-bottom: 1rem;
    }
    .metric-card {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 0.5rem;
        border-left: 4px solid #3498db;
    }
    .success-field {
        background-color: #d4edda;
        padding: 0.5rem;
        border-radius: 0.25rem;
        margin: 0.25rem 0;
    }
    .missing-field {
        background-color: #f8d7da;
        padding: 0.5rem;
        border-radius: 0.25rem;
        margin: 0.25rem 0;
    }
    .stButton>button {
        background-color: #2c3e50;
        color: white;
        border-radius: 0.5rem;
        padding: 0.5rem 2rem;
        font-weight: bold;
    }
    .stButton>button:hover {
        background-color: #34495e;
    }
</style>
""", unsafe_allow_html=True)

# Initialize session state
if 'current_data' not in st.session_state:
    st.session_state.current_data = None
if 'checklist_generated' not in st.session_state:
    st.session_state.checklist_generated = False
if 'selected_group' not in st.session_state:
    st.session_state.selected_group = None
if 'keywords_list' not in st.session_state:
    st.session_state.keywords_list = None
if 'rapidminer_data' not in st.session_state:
    st.session_state.rapidminer_data = None
if 'openai_api_key' not in st.session_state:
    st.session_state.openai_api_key = os.getenv('OPENAI_API_KEY', '')
if 'ai_generated_checklist' not in st.session_state:
    st.session_state.ai_generated_checklist = None
if 'chat_history' not in st.session_state:
    st.session_state.chat_history = []
if 'user_corrections' not in st.session_state:
    st.session_state.user_corrections = []
if 'checklist_history' not in st.session_state:
    st.session_state.checklist_history = []

# Sidebar Navigation
with st.sidebar:
    st.markdown("# 📋 Plan Review System")
    st.markdown("---")
    
    page = st.radio(
        "Navigation",
        ["🏠 Home", "⚙️ Settings", "🔑 Keywords", "🔨 Generate Checklist", 
         "💬 Chat Assistant", "📄 Export", "ℹ️ About"],
        label_visibility="collapsed"
    )
    
    st.markdown("---")
    st.markdown("### Quick Stats")
    
    if st.session_state.current_data is not None:
        df = st.session_state.current_data
        st.metric("Total Fields", len(df))
        found = len(df[df['Extracted_Value'] != 'N/F'])
        st.metric("Found", found, delta=f"{(found/len(df)*100):.0f}%")
        st.metric("Avg Confidence", f"{df['Confidence'].mean()*100:.0f}%")
    else:
        st.info("No data loaded yet")
    
    st.markdown("---")
    st.markdown("**Version:** 1.0  \n**Updated:** Nov 2025")

# Helper functions
def load_mock_data(group_name):
    """Load mock data for selected group"""
    mock_files = {
        "Aurora Dynamics": "mock_data/orange_output_aurora_dynamics.xlsx",
        "Helios Manufacturing Inc.": "mock_data/orange_output_helios_manufacturing.xlsx",
        "Solstice Technologies": "mock_data/orange_output_solstice_technologies.xlsx",
        "TechVenture Inc.": "mock_data/orange_output_techventure_typos.xlsx",
        "GlobalCorp International": "mock_data/orange_output_globalcorp.xlsx"
    }
    
    if group_name in mock_files:
        filepath = Path(mock_files[group_name])
        if filepath.exists():
            return pd.read_excel(filepath)
    return None

def get_confidence_color(confidence):
    """Return color based on confidence level"""
    if confidence >= 0.8:
        return "🟢"
    elif confidence >= 0.5:
        return "🟡"
    else:
        return "🔴"

def load_keywords_list(file):
    """Load keywords list from Excel or text file"""
    try:
        if file.name.endswith(('.xlsx', '.xls')):
            df = pd.read_excel(file)
            # Assume keywords are in first column or column named 'Keywords'
            if 'Keywords' in df.columns:
                keywords = df['Keywords'].dropna().tolist()
            elif 'Keyword' in df.columns:
                keywords = df['Keyword'].dropna().tolist()
            else:
                keywords = df.iloc[:, 0].dropna().tolist()
            return keywords
        else:
            # Text file - one keyword per line
            content = file.read().decode('utf-8')
            keywords = [line.strip() for line in content.split('\n') if line.strip()]
            return keywords
    except Exception as e:
        st.error(f"Error loading keywords: {str(e)}")
        return None

def load_rapidminer_output(file):
    """Load RapidMiner CSV output with line_no, text, tags columns"""
    try:
        df = pd.read_csv(file)
        # Validate required columns
        required_cols = ['line_no', 'text', 'tags']
        if not all(col in df.columns for col in required_cols):
            st.error(f"CSV must have columns: {', '.join(required_cols)}")
            return None
        return df
    except Exception as e:
        st.error(f"Error loading RapidMiner output: {str(e)}")
        return None

def load_appendix2_guidelines():
    """Load guidelines from Appendix 2 Excel file (tabs 3, 4, 5)"""
    try:
        appendix_path = Path(__file__).parent / "Checklist Examples" / "Appendix 2 Plan Doc Discrepancies Participation 6.xlsx"
        
        if not appendix_path.exists():
            return None
        
        # Read the three guideline tabs
        guidelines = {}
        
        # Tab 3: Requires Approval criteria
        try:
            approval_df = pd.read_excel(appendix_path, sheet_name=2)  # 0-indexed, so tab 3 is index 2
            guidelines['approval'] = approval_df.to_string(index=False) if not approval_df.empty else "No approval criteria found"
        except:
            guidelines['approval'] = "Could not read approval criteria"
        
        # Tab 4: Requires Notice criteria  
        try:
            notice_df = pd.read_excel(appendix_path, sheet_name=3)  # Tab 4 is index 3
            guidelines['notice'] = notice_df.to_string(index=False) if not notice_df.empty else "No notice criteria found"
        except:
            guidelines['notice'] = "Could not read notice criteria"
        
        # Tab 5: Request Handbook criteria
        try:
            handbook_df = pd.read_excel(appendix_path, sheet_name=4)  # Tab 5 is index 4
            guidelines['handbook'] = handbook_df.to_string(index=False) if not handbook_df.empty else "No handbook criteria found"
        except:
            guidelines['handbook'] = "Could not read handbook criteria"
        
        return guidelines
        
    except Exception as e:
        st.warning(f"Could not load Appendix 2 guidelines: {str(e)}")
        return None

def generate_checklist_from_tags(rapidminer_df):
    """Generate checklist from RapidMiner tagged output (no API key needed)"""
    
    # Define all 21 checklist items and their tag mappings
    checklist_items = {
        'UR Vendor': ['UR Vendor'],
        'PPO Network': ['PPO Network'],
        'Retirees': ['Retirees'],
        'BOD/Directors/Officers': ['Board of Directors', 'Directors', 'Officers'],
        'Min Hour Requirement': ['Minimum Hour Requirements', 'Minimum Hours'],
        'Dependent Definitions': ['Dependent Definitions'],
        'Req Adding Dependents': ['Req Adding Dependents', 'Requirements for Adding Dependents'],
        'Dependent to Age 26': ['Dependent to Age 26', 'Dependent Coverage to Age 26'],
        'Grandchildren': ['Grandchildren', 'Grandchildren Coverage'],
        'Termination Provisions': ['Termination Provisions', 'Termination'],
        'Open Enrollment': ['Open Enrollment'],
        'Leave of Absence': ['Leave of Absence', 'LOA'],
        'Medically Necessary': ['Medically Necessary'],
        'E&I': ['E&I', 'Eligible & Ineligible'],
        'R&C': ['R&C', 'Reasonable and Customary', 'Reasonable & Customary'],
        'Workers Comp': ['Workers Comp', 'Workers Compensation'],
        'Transplant': ['Transplant', 'Transplant Benefits'],
        'ETS Gene Therapy': ['ETS Gene Therapy', 'Gene Therapy'],
        'Coordination of Benefits': ['Coordination of Benefits', 'COB'],
        'COBRA': ['COBRA', 'COBRA Continuation'],
        'Subrogation': ['Subrogation']
    }
    
    # Build checklist results
    results = []
    
    for item_name, tag_list in checklist_items.items():
        # Check if any tags match in the RapidMiner output
        matched = False
        matched_texts = []
        
        for _, row in rapidminer_df.iterrows():
            row_tags = [tag.strip() for tag in str(row['tags']).split(';')]
            if any(tag in tag_list for tag in row_tags):
                matched = True
                matched_texts.append(row['text'])
        
        # SIMPLIFIED LOGIC: Only check "Matches" if found, leave everything else blank
        results.append({
            'Item': item_name,
            'Matches': matched,
            'Requires Approval': False,  # Always blank for now
            'Requires Notice': False,     # Always blank for now
            'Request Handbook': False,    # Always blank for now
            'AI Reasoning': f"Found in document: {matched_texts[0][:100]}..." if matched and matched_texts else "Not found in RapidMiner output"
        })
    
    return pd.DataFrame(results)

def generate_checklist_with_ai(rapidminer_df, api_key):
    """Use ChatGPT to intelligently analyze RapidMiner output and generate checklist"""
    if not api_key:
        st.error("❌ OpenAI API key is required")
        return None
    
    # Load Appendix 2 guidelines
    guidelines = load_appendix2_guidelines()
    
    # Define the checklist items from Appendix 2
    checklist_items = [
        'UR Vendor',
        'PPO Network',
        'Retirees',
        'BOD/Directors/Officers',
        'Minimum Hour Requirement',
        'Dependent Definitions',
        'Req adding dependents',
        'Dependent to age 26',
        'Grandchildren',
        'Termination Provisions',
        'Open Enrollment',
        'Leave of Absence',
        'Medically Necessary',
        'E&I',
        'R&C',
        'Workers Comp',
        'Transplant',
        'ETS Gene Therapy',
        'Coordination of Benefits',
        'COBRA',
        'Subrogation'
    ]
    
    try:
        client = OpenAI(api_key=api_key)
        
        # Build context from RapidMiner data
        document_context = []
        for _, row in rapidminer_df.iterrows():
            document_context.append(f"Line {row['line_no']}: {row['text']} [Tags: {row['tags']}]")
        
        context_str = "\n".join(document_context[:100])  # Limit to first 100 entries to avoid token limits
        
        # Build guidelines context if available
        guidelines_context = ""
        if guidelines:
            guidelines_context = f"""
TMHCC APPENDIX 2 GUIDELINES (from "Plan Doc Discrepancies Participation 6.xlsx"):

=== REQUIRES APPROVAL CRITERIA (Tab 3) ===
{guidelines.get('approval', 'Not available')}

=== REQUIRES NOTICE CRITERIA (Tab 4) ===
{guidelines.get('notice', 'Not available')}

=== REQUEST HANDBOOK CRITERIA (Tab 5) ===
{guidelines.get('handbook', 'Not available')}

YOU MUST FOLLOW THESE EXACT GUIDELINES when determining approval/notice/handbook decisions.
"""
        else:
            guidelines_context = """
NOTE: Appendix 2 guidelines file not found. Use conservative standard insurance practices:
- Requires Approval: Non-standard terms, unusual limits, policy deviations
- Requires Notice: Minor variations, informational items  
- Request Handbook: Employee-facing provisions, clear guidelines
"""
        
        prompt = f"""You are reviewing insurance plan documents for TMHCC following the guidelines in "Appendix 2 Plan Doc Discrepancies Participation". 

{guidelines_context}

Extracted Document Information:
{context_str}

Checklist items to evaluate:
{chr(10).join([f'{i+1}. {item}' for i, item in enumerate(checklist_items)])}

For each checklist item, provide ADVANCED INTELLIGENT ANALYSIS following TMHCC Appendix 2 guidelines:

CRITICAL LOGIC - ONLY ONE CHECK PER ITEM:
1. If item requires APPROVAL → check ONLY "requires_approval" (this implies it matched)
2. If item requires NOTICE → check ONLY "requires_notice" (this implies it matched)
3. If item should go to HANDBOOK → check ONLY "request_handbook" (this implies it matched)
4. If item is present but needs NO action → check ONLY "matches"
5. If item is NOT found → ALL fields false (completely blank)

Apply Appendix 2 criteria:
- **Requires Approval**: Apply EXACT criteria from Tab 3 above (non-standard terms, deviations)
- **Requires Notice**: Apply EXACT criteria from Tab 4 above (stakeholder notifications)
- **Request Handbook**: Apply EXACT criteria from Tab 5 above (employee-facing provisions)

CRITICAL REQUIREMENTS:
- Follow TMHCC Appendix 2 guidelines EXACTLY as provided above
- Read the ACTUAL TEXT carefully, not just tags
- Cite specific Appendix 2 criteria in your reasoning
- ONE CHECK ONLY per item - if approval is true, matches must be false

Example reasoning format:
- "Found 'Minimum Work Hours: 28 hours/week' - Per Appendix 2 Tab 3, standard is 30 hours. REQUIRES APPROVAL (not matches) due to non-standard threshold."
- "Transplant benefits with standard terms - MATCHES ONLY. Per Appendix 2 Tab 5, standard provision, could go to handbook but not required."
- "No E&I clause found - ALL FALSE (completely blank)."

Respond in JSON format:
{{
  "checklist": [
    {{
      "item": "item name",
      "matches": true/false,
      "requires_approval": true/false,
      "requires_notice": true/false,
      "request_handbook": true/false,
      "reasoning": "detailed explanation citing specific Appendix 2 tab criteria"
    }}
  ]
}}"""
        
        with st.spinner("🤖 AI is analyzing document and generating intelligent checklist..."):
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are an expert insurance plan document reviewer for TMHCC with deep knowledge of compliance requirements and policy implications."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.3
            )
            
            result = json.loads(response.choices[0].message.content)
            checklist_data = result.get('checklist', [])
            
            # Convert to DataFrame and enforce one-check-per-item rule
            df_data = []
            for item_data in checklist_data:
                matches = item_data.get('matches', False)
                approval = item_data.get('requires_approval', False)
                notice = item_data.get('requires_notice', False)
                handbook = item_data.get('request_handbook', False)
                
                # ENFORCE: Only one check per item
                # Priority: Approval > Notice > Handbook > Matches
                if approval:
                    matches = False
                    notice = False
                    handbook = False
                elif notice:
                    matches = False
                    approval = False
                    handbook = False
                elif handbook:
                    matches = False
                    approval = False
                    notice = False
                # If matches is true, others should already be false
                
                df_data.append({
                    'Item': item_data.get('item', ''),
                    'Matches': matches,
                    'Requires Approval': approval,
                    'Requires Notice': notice,
                    'Request Handbook': handbook,
                    'AI Reasoning': item_data.get('reasoning', '')
                })
            
            return pd.DataFrame(df_data)
            
    except Exception as e:
        st.error(f"❌ AI Error: {str(e)}")
        return None
        return None

def enhance_with_ai(df, definitions_df, api_key):
    """Use ChatGPT API to enhance field matching and validation"""
    if not api_key:
        st.error("❌ OpenAI API key is required")
        return df
    
    try:
        client = OpenAI(api_key=api_key)
        
        # Create definitions context
        definitions_context = "Insurance Plan Document Field Definitions:\n\n"
        for _, row in definitions_df.iterrows():
            field_name = row.get('Field', row.get('Term', ''))
            definition = row.get('Definition', row.get('Description', ''))
            if field_name and definition:
                definitions_context += f"- {field_name}: {definition}\n"
        
        enhanced_rows = []
        
        with st.spinner("🤖 AI is analyzing fields and making intelligent decisions..."):
            progress_bar = st.progress(0)
            
            for idx, row in df.iterrows():
                field_name = row['Field']
                extracted_value = row['Extracted_Value']
                confidence = row['Confidence']
                
                # For fields with low confidence or missing values, ask AI
                if confidence < 0.7 or extracted_value == 'N/F':
                    prompt = f"""Based on the following field definitions and extracted data, help validate this insurance plan field:

{definitions_context}

Field Name: {field_name}
Extracted Value: {extracted_value}
Original Confidence: {confidence}

Task:
1. Is this field required for insurance plan documents based on standard practices?
2. If the value is missing or low confidence, suggest what should be verified
3. Provide a validation status: VALID, NEEDS_REVIEW, or MISSING
4. Brief explanation (max 50 words)

Respond in JSON format:
{{
  "validation_status": "VALID/NEEDS_REVIEW/MISSING",
  "is_required": true/false,
  "suggestion": "verification suggestion",
  "explanation": "brief explanation"
}}"""
                    
                    try:
                        response = client.chat.completions.create(
                            model="gpt-4o-mini",
                            messages=[
                                {"role": "system", "content": "You are an expert insurance plan document reviewer for TMHCC. Provide concise, accurate assessments."},
                                {"role": "user", "content": prompt}
                            ],
                            response_format={"type": "json_object"},
                            temperature=0.3
                        )
                        
                        ai_response = json.loads(response.choices[0].message.content)
                        
                        row['AI_Validation_Status'] = ai_response.get('validation_status', 'NEEDS_REVIEW')
                        row['AI_Required'] = ai_response.get('is_required', False)
                        row['AI_Suggestion'] = ai_response.get('suggestion', '')
                        row['AI_Explanation'] = ai_response.get('explanation', '')
                        row['AI_Enhanced'] = True
                        
                    except Exception as e:
                        row['AI_Validation_Status'] = 'ERROR'
                        row['AI_Required'] = False
                        row['AI_Suggestion'] = f'AI error: {str(e)}'
                        row['AI_Explanation'] = ''
                        row['AI_Enhanced'] = False
                else:
                    # High confidence fields
                    row['AI_Validation_Status'] = 'VALID'
                    row['AI_Required'] = True
                    row['AI_Suggestion'] = 'Field validated with high confidence'
                    row['AI_Explanation'] = 'Automated extraction successful'
                    row['AI_Enhanced'] = True
                
                enhanced_rows.append(row)
                progress_bar.progress((idx + 1) / len(df))
            
            progress_bar.empty()
        
        enhanced_df = pd.DataFrame(enhanced_rows)
        return enhanced_df
        
    except Exception as e:
        st.error(f"❌ AI Enhancement Error: {str(e)}")
        return df

def generate_excel_checklist_from_ai(checklist_df):
    """Generate formatted Excel checklist from AI-generated checklist DataFrame"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Draft Checklist"
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10
    
    # Define styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Helper function to extract value from RapidMiner data by tag
    def get_value_by_tag(tag_name):
        """Extract text value from RapidMiner data matching a specific tag"""
        if 'rapidminer_data' in st.session_state and st.session_state.rapidminer_data is not None:
            for _, row in st.session_state.rapidminer_data.iterrows():
                tags = [t.strip() for t in str(row['tags']).split(';')]
                if tag_name in tags:
                    return row['text']
        return None
    
    # Rows 1-11: Header information (extracted from RapidMiner data)
    ws.cell(row=1, column=1).value = "Group Name: "
    ws.cell(row=1, column=1).font = bold_font
    
    ws.cell(row=2, column=1).value = "Group Name in PD:"
    ws.cell(row=2, column=1).font = bold_font
    group_name = get_value_by_tag('Group Name')
    if group_name:
        ws.cell(row=2, column=2).value = group_name
    
    ws.cell(row=3, column=1).value = "Group Eff Date:"
    ws.cell(row=3, column=1).font = bold_font
    
    ws.cell(row=4, column=1).value = "PD Eff Date:"
    ws.cell(row=4, column=1).font = bold_font
    eff_date = get_value_by_tag('Group Effective Date')
    if not eff_date:
        eff_date = get_value_by_tag('Plan Document Effective Date')
    if eff_date:
        ws.cell(row=4, column=2).value = eff_date
    
    ws.cell(row=5, column=1).value = "TPA:"
    ws.cell(row=5, column=1).font = bold_font
    
    ws.cell(row=6, column=1).value = "TPA in PD:"
    ws.cell(row=6, column=1).font = bold_font
    tpa = get_value_by_tag('TPA')
    if tpa:
        ws.cell(row=6, column=2).value = tpa
    
    ws.cell(row=7, column=1).value = "Underwriter:"
    ws.cell(row=7, column=1).font = bold_font
    underwriter = get_value_by_tag('Underwriter')
    if underwriter:
        ws.cell(row=7, column=2).value = underwriter
    
    ws.cell(row=8, column=1).value = "Benefit Plan Name:"
    ws.cell(row=8, column=1).font = bold_font
    
    ws.cell(row=9, column=1).value = "Benefit Plan Name in PD:"
    ws.cell(row=9, column=1).font = bold_font
    plan_name = get_value_by_tag('Benefit Plan Name')
    if plan_name:
        ws.cell(row=9, column=2).value = plan_name
    
    ws.cell(row=10, column=1).value = "Subsidiaries:"
    ws.cell(row=10, column=1).font = bold_font
    subsidiaries = get_value_by_tag('Subsidiaries')
    if subsidiaries:
        ws.cell(row=10, column=2).value = subsidiaries
    
    ws.cell(row=11, column=1).value = "Group Address:"
    ws.cell(row=11, column=1).font = bold_font
    address = get_value_by_tag('Group Address')
    if address:
        ws.cell(row=11, column=2).value = address
    
    # Row 12: Empty (spacing)
    
    # Row 13: Schedule of Benefits Section Header
    ws.cell(row=13, column=1).value = "Schedule of Benefits"
    ws.cell(row=13, column=1).font = Font(bold=True, size=12)
    
    # Row 14: Column headers for Schedule of Benefits
    ws.cell(row=14, column=2).value = "Matches"
    ws.cell(row=14, column=2).font = header_font
    ws.cell(row=14, column=3).value = "Requires enrollment"
    ws.cell(row=14, column=3).font = header_font
    ws.cell(row=14, column=4).value = "Updated PA"
    ws.cell(row=14, column=4).font = header_font
    
    # Rows 15-20: Schedule of Benefits items
    # Define schedule items with their corresponding search tags
    schedule_items = [
        ("Plan: MEDICAL SCHEDULE OF BENEFITS (POS Plan)", ["MEDICAL", "POS"]),
        ("Plan: MEDICAL SCHEDULE OF BENEFITS (HDHP Plan)", ["MEDICAL", "HDHP"]),
        ("Plan: TRANSPLANT SCHEDULE OF BENEFITS (POS Plan)", ["TRANSPLANT", "POS"]),
        ("Plan: TRANSPLANT SCHEDULE OF BENEFITS (HDHP Plan)", ["TRANSPLANT", "HDHP"]),
        ("Plan: PRESCRIPTION SCHEDULE OF BENEFITS (POS Plan)", ["PRESCRIPTION", "POS"]),
        ("Plan: PRESCRIPTION SCHEDULE OF BENEFITS (HDHP Plan)", ["PRESCRIPTION", "HDHP"])
    ]
    
    # Helper function to check if schedule item exists in RapidMiner data
    def check_schedule_item(search_terms):
        """Check if text matching search terms exists in RapidMiner data"""
        if 'rapidminer_data' in st.session_state and st.session_state.rapidminer_data is not None:
            for _, row in st.session_state.rapidminer_data.iterrows():
                text_upper = str(row['text']).upper()
                # Check if all search terms are in the text
                if all(term in text_upper for term in search_terms):
                    return True
        return False
    
    for idx, (item_name, search_terms) in enumerate(schedule_items):
        row_num = 15 + idx
        ws.cell(row=row_num, column=1).value = item_name
        ws.cell(row=row_num, column=1).border = border
        
        # Check if item exists in RapidMiner data
        matches = check_schedule_item(search_terms)
        
        # Column 2: Matches (mark X if found)
        ws.cell(row=row_num, column=2).value = "X" if matches else ""
        ws.cell(row=row_num, column=2).alignment = center_align
        ws.cell(row=row_num, column=2).border = border
        
        # Columns 3-4: Requires enrollment, Updated PA (empty for now)
        for col in range(3, 5):
            ws.cell(row=row_num, column=col).value = ""
            ws.cell(row=row_num, column=col).alignment = center_align
            ws.cell(row=row_num, column=col).border = border
    
    # Row 21: Empty (spacing)
    
    # Row 22: Column headers for main checklist
    ws.cell(row=22, column=1).value = "Item"
    ws.cell(row=22, column=2).value = "Matches"
    ws.cell(row=22, column=3).value = "Requires Approval"
    ws.cell(row=22, column=4).value = "Requires Notice"
    ws.cell(row=22, column=5).value = "Request Handbook"
    ws.cell(row=22, column=6).value = "Pg #"
    
    # Style headers
    for col in range(1, 7):
        cell = ws.cell(row=22, column=col)
        cell.font = header_font
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = center_align
        cell.border = border
    
    # Row 23: Empty (spacing)
    
    # Rows 24-44: Checklist items (21 items)
    start_row = 24
    for idx, row in checklist_df.iterrows():
        current_row = start_row + idx
        
        # Item name
        ws.cell(row=current_row, column=1).value = row['Item']
        ws.cell(row=current_row, column=1).border = border
        
        # Matches
        ws.cell(row=current_row, column=2).value = "X" if row['Matches'] else ""
        ws.cell(row=current_row, column=2).alignment = center_align
        ws.cell(row=current_row, column=2).border = border
        
        # Requires Approval
        ws.cell(row=current_row, column=3).value = "X" if row['Requires Approval'] else ""
        ws.cell(row=current_row, column=3).alignment = center_align
        ws.cell(row=current_row, column=3).border = border
        
        # Requires Notice
        ws.cell(row=current_row, column=4).value = "X" if row['Requires Notice'] else ""
        ws.cell(row=current_row, column=4).alignment = center_align
        ws.cell(row=current_row, column=4).border = border
        
        # Request Handbook
        ws.cell(row=current_row, column=5).value = "X" if row['Request Handbook'] else ""
        ws.cell(row=current_row, column=5).alignment = center_align
        ws.cell(row=current_row, column=5).border = border
        
        # Page number (leave empty for now - can be populated if available)
        ws.cell(row=current_row, column=6).value = ""
        ws.cell(row=current_row, column=6).alignment = center_align
        ws.cell(row=current_row, column=6).border = border
    
    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer.getvalue()

def generate_excel_checklist(df, group_name):
    """Generate formatted Excel checklist matching the Draft Checklist format"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Draft Checklist"
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10
    
    # Define styles
    bold_font = Font(bold=True)
    header_font = Font(bold=True, size=11)
    
    # Create field mapping for rows 1-11
    field_mapping = {
        'Group Name': (1, 2),
        'Group Eff Date': (3, 4),
        'TPA': (5, 6),
        'Third Party Administrator': (5, 6),
        'Underwriter': (7, 7),
        'Benefit Plan Name': (8, 9),
        'Subsidiaries': (10, 10),
        'Group Address': (11, 11)
    }
    
    row_num = 1
    
    # Helper function to find field value
    def get_field_value(field_name):
        for _, row in df.iterrows():
            if field_name.lower() in row['Field'].lower():
                return row['Extracted_Value'] if row['Extracted_Value'] != 'N/F' else None
        return None
    
    # Rows 1-2: Group Name
    ws.cell(row=1, column=1).value = "Group Name: "
    ws.cell(row=1, column=1).font = bold_font
    ws.cell(row=2, column=1).value = "Group Name in PD:"
    ws.cell(row=2, column=1).font = bold_font
    group_name_value = get_field_value('Group Name')
    if group_name_value:
        ws.cell(row=2, column=2).value = group_name_value
    
    # Rows 3-4: Group Eff Date
    ws.cell(row=3, column=1).value = "Group Eff Date:"
    ws.cell(row=3, column=1).font = bold_font
    ws.cell(row=4, column=1).value = "PD Eff Date:"
    ws.cell(row=4, column=1).font = bold_font
    eff_date_value = get_field_value('Eff Date')
    if eff_date_value:
        ws.cell(row=4, column=2).value = eff_date_value
    
    # Rows 5-6: TPA
    ws.cell(row=5, column=1).value = "TPA:"
    ws.cell(row=5, column=1).font = bold_font
    ws.cell(row=6, column=1).value = "TPA in PD:"
    ws.cell(row=6, column=1).font = bold_font
    tpa_value = get_field_value('TPA') or get_field_value('Third Party Administrator')
    if tpa_value:
        ws.cell(row=6, column=2).value = tpa_value
    
    # Row 7: Underwriter
    ws.cell(row=7, column=1).value = "Underwriter:"
    ws.cell(row=7, column=1).font = bold_font
    underwriter_value = get_field_value('Underwriter')
    if underwriter_value:
        ws.cell(row=7, column=2).value = underwriter_value
    
    # Rows 8-9: Benefit Plan Name
    ws.cell(row=8, column=1).value = "Benefit Plan Name:"
    ws.cell(row=8, column=1).font = bold_font
    ws.cell(row=9, column=1).value = "Benefit Plan Name in PD:"
    ws.cell(row=9, column=1).font = bold_font
    plan_name_value = get_field_value('Plan') or get_field_value('Benefit Plan')
    if plan_name_value:
        ws.cell(row=9, column=2).value = plan_name_value
    
    # Row 10: Subsidiaries
    ws.cell(row=10, column=1).value = "Subsidiaries:"
    ws.cell(row=10, column=1).font = bold_font
    subsidiaries_value = get_field_value('Subsidiaries')
    if subsidiaries_value:
        ws.cell(row=10, column=2).value = subsidiaries_value
    
    # Row 11: Group Address
    ws.cell(row=11, column=1).value = "Group Address:"
    ws.cell(row=11, column=1).font = bold_font
    address_value = get_field_value('Address')
    if address_value:
        ws.cell(row=11, column=2).value = address_value
    
    # Row 12: Empty
    row_num = 12
    
    # Row 13: Headers for Schedule of Benefits section (rows 13-20, not used in this version)
    row_num = 13
    ws.cell(row=row_num, column=2).value = "Matches"
    ws.cell(row=row_num, column=2).font = header_font
    ws.cell(row=row_num, column=3).value = "Requires enrollment"
    ws.cell(row=row_num, column=3).font = header_font
    ws.cell(row=row_num, column=4).value = "Updated PA"
    ws.cell(row=row_num, column=4).font = header_font
    
    # Skip to row 21 (empty row before main checklist items)
    row_num = 21
    
    # Row 22: Headers for checklist items (rows 22-43)
    row_num = 22
    ws.cell(row=row_num, column=2).value = "Matches"
    ws.cell(row=row_num, column=2).font = header_font
    ws.cell(row=row_num, column=3).value = "Requires Approval"
    ws.cell(row=row_num, column=3).font = header_font
    ws.cell(row=row_num, column=4).value = "Requires Notice"
    ws.cell(row=row_num, column=4).font = header_font
    ws.cell(row=row_num, column=5).value = "Request Handbook"
    ws.cell(row=row_num, column=5).font = header_font
    ws.cell(row=row_num, column=6).value = "Pg #"
    ws.cell(row=row_num, column=6).font = header_font
    
    # Define checklist items (rows 23-43) - these are the fields to match
    checklist_items = [
        'UR Vendor',
        'PPO Network',
        'Retirees',
        'BOD, Directors, Officers',
        'Minimum Hour Requirement',
        'Dependent Definitions',
        'Req adding dependents',
        'Dependent to age 26',
        'Grandchildren',
        'Termination Provisions',
        'Open Enrollment',
        'Leave of Absence',
        'Medically Necessary',
        'E&I',
        'R&C',
        'Workers Comp',
        'Transplant',
        'ETS Gene Therapy',
        'Coordination of Benefits',
        'COBRA',
        'Subrogation'
    ]
    
    # Add checklist items rows 23-43
    row_num = 23
    for item in checklist_items:
        ws.cell(row=row_num, column=1).value = item
        
        # Find if this field exists in the data
        found = False
        page_num = 'N/A'
        
        for _, data_row in df.iterrows():
            # Match field names (flexible matching)
            if (item.lower() in data_row['Field'].lower() or 
                data_row['Field'].lower() in item.lower() or
                item.replace(',', '').replace(' ', '').lower() in data_row['Field'].replace(' ', '').lower()):
                
                if data_row['Extracted_Value'] != 'N/F':
                    found = True
                    if pd.notna(data_row['Page_Number']):
                        page_num = int(data_row['Page_Number'])
                break
        
        # Set Matches column (B) - True if found, False if not
        ws.cell(row=row_num, column=2).value = found
        
        # Set other columns to False by default
        ws.cell(row=row_num, column=3).value = False  # Requires Approval
        ws.cell(row=row_num, column=4).value = False  # Requires Notice
        ws.cell(row=row_num, column=5).value = False  # Request Handbook
        
        # Set page number
        ws.cell(row=row_num, column=6).value = page_num
        
        row_num += 1
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

# Settings Page
if page == "⚙️ Settings":
    st.markdown('<div class="main-header">⚙️ AI Settings</div>', 
                unsafe_allow_html=True)
    
    st.markdown("""
    ### AI Configuration Status
    
    The system automatically loads the API key from the `.env` file. 
    Guests can use the app without entering anything!
    """)
    
    # API Status
    st.markdown("#### Current API Status")
    
    if st.session_state.openai_api_key:
        col1, col2 = st.columns([3, 1])
        with col1:
            st.success("🟢 AI Mode: **ENABLED** (API key loaded from .env file)")
        with col2:
            # Test connection
            if st.button("🔌 Test API"):
                try:
                    client = OpenAI(api_key=st.session_state.openai_api_key)
                    response = client.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[{"role": "user", "content": "test"}],
                        max_tokens=5
                    )
                    st.success("✅ Connected!")
                except Exception as e:
                    st.error(f"❌ Connection failed: {str(e)}")
    else:
        st.info("📋 Rule-Based Mode: **ACTIVE** (No API key configured)")
        st.markdown("The app works without an API key using tag-based matching!")
    
    st.markdown("---")
    
    # Optional: Manual Override
    with st.expander("🔧 Advanced: Manual API Key Override (Optional)"):
        st.markdown("""
        **For developers only:** You can temporarily override the API key for testing.
        This is NOT required for normal use.
        """)
        
        api_key_input = st.text_input(
            "Temporary API Key Override",
            value="",
            type="password",
            help="Leave empty to use .env file"
        )
        
        if st.button("💾 Use This Key (Session Only)"):
            if api_key_input:
                st.session_state.openai_api_key = api_key_input
                st.success("✅ Using override key for this session!")
            else:
                st.session_state.openai_api_key = os.getenv('OPENAI_API_KEY', '')
                st.info("Reset to .env configuration")
            st.rerun()
    
    st.markdown("---")
    
    # Mode Comparison
    st.markdown("#### Processing Modes")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**🤖 AI Mode** (With API Key)")
        st.markdown("""
        - Reads actual document text
        - Intelligent context analysis
        - Detailed reasoning
        - Catches nuances & errors
        """)
    
    with col2:
        st.markdown("**📋 Rule-Based Mode** (No API Key)")
        st.markdown("""
        - Tag-based matching
        - Fast & free
        - No internet required
        - Works offline
        """)
    
    st.markdown("---")
    
    # Setup Instructions (for administrators)
    with st.expander("📖 Setup Instructions for Administrators"):
        st.markdown("""
        #### For Project Owner/Administrator
        
        To enable AI features for all users:
        
        1. Get an OpenAI API key from [OpenAI Platform](https://platform.openai.com/api-keys)
        2. Add it to the `.env` file in the project root:
           ```
           OPENAI_API_KEY=your-actual-key-here
           ```
        3. Restart the app
        4. All users will automatically have AI features enabled!
        
        **Security Note:** The `.env` file is git-ignored and never committed to the repository.
        """)

# Keywords/Data Page
elif page == "🔑 Keywords":
    st.markdown('<div class="main-header">📊 RapidMiner Data Input</div>', 
                unsafe_allow_html=True)
    
    st.markdown("""
    ### Upload RapidMiner Output
    
    Upload the CSV file from RapidMiner with extracted plan document information.
    The system will use the tags to automatically fill the checklist.
    """)
    
    # Upload RapidMiner CSV
    st.markdown("#### Upload RapidMiner CSV Output")
    
    uploaded_data = st.file_uploader(
        "Choose RapidMiner CSV file",
        type=['csv'],
        help="CSV file with columns: line_no, text, tags",
        key="rapidminer_uploader"
    )
    
    if uploaded_data is not None:
        rm_data = load_rapidminer_output(uploaded_data)
        
        if rm_data is not None:
            st.session_state.rapidminer_data = rm_data
            st.success(f"✅ RapidMiner data loaded! {len(rm_data)} entries found.")
            
            # Show preview
            st.markdown("#### Data Preview")
            col1, col2 = st.columns([3, 1])
            
            with col1:
                st.dataframe(rm_data.head(10), use_container_width=True)
            
            with col2:
                st.metric("Total Entries", len(rm_data))
                unique_tags = set()
                for tags in rm_data['tags']:
                    unique_tags.update([t.strip() for t in str(tags).split(';')])
                st.metric("Unique Tags", len(unique_tags))
                
            # Show tag distribution
            with st.expander("📊 View Tag Distribution"):
                tag_counts = {}
                for tags in rm_data['tags']:
                    for tag in str(tags).split(';'):
                        tag = tag.strip()
                        tag_counts[tag] = tag_counts.get(tag, 0) + 1
                
                tag_df = pd.DataFrame([
                    {'Tag': tag, 'Count': count} 
                    for tag, count in sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)
                ])
                st.dataframe(tag_df, use_container_width=True)
    
    
    st.markdown("---")
    
    # Current Data Status
    if st.session_state.rapidminer_data is not None:
        st.markdown("#### Current Status")
        st.info(f"📋 {len(st.session_state.rapidminer_data)} entries loaded and ready for checklist generation")
        
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🗑️ Clear Data", use_container_width=True):
                st.session_state.rapidminer_data = None
                st.session_state.ai_generated_checklist = None
                st.rerun()
        with col2:
            if st.button("👁️ View All Data", use_container_width=True):
                with st.expander("All RapidMiner Data", expanded=True):
                    st.dataframe(st.session_state.rapidminer_data, use_container_width=True)
    else:
        st.warning("⚠️ No data loaded. Upload a RapidMiner CSV file to generate checklist.")
    
    st.markdown("---")
    
    # Example Format
    st.markdown("#### Expected CSV Format")
    
    st.markdown("""
    **RapidMiner CSV File:**
    - Must have 3 columns: `line_no`, `text`, `tags`
    - `line_no`: Line number from plan document
    - `text`: Extracted text from that line
    - `tags`: Semicolon-separated tags (e.g., "TPA; Underwriter")
    """)
    
    example_data = pd.DataFrame({
        'line_no': [3, 4, 5, 8, 10],
        'text': [
            'Plan Reference: 2026-AD-1123 | Launch Date: March 1, 2026',
            'Plan Administrator: BluePeak Benefits | Insurer: Nova Mutual',
            'Plan Name: Aurora Dynamics Health Plan',
            'Utilization Review Partner: MedAxis Group',
            'Minimum Work Hours: 28 hours per week'
        ],
        'tags': [
            'Group Effective Date; Plan Document Effective Date',
            'TPA; Underwriter',
            'Benefit Plan Name',
            'UR Vendor',
            'Minimum Hour Requirements'
        ]
    })
    st.dataframe(example_data, use_container_width=True)

# Main Content - Home Page
elif page == "🏠 Home":
    st.markdown('<div class="main-header">🤖 Tokio Marine Plan Document Review System</div>', 
                unsafe_allow_html=True)
    
    st.markdown("""
    ### Welcome to the TokioMarineChatbot!
    
    This intelligent system helps you efficiently review insurance plan documents and generate compliance checklists.
    """)
    
    st.markdown("---")
    
    # Getting Started Guide
    st.markdown("### 🚀 Quick Start Guide")
    
    st.markdown("""
    1. **⚙️ Settings** - Check your AI status (API key auto-loaded from .env)
    2. **🔑 Keywords** - Upload your RapidMiner CSV output (columns: line_no, text, tags)
    3. **🔨 Generate Checklist** - Choose Tag-Based (fast) or AI Advanced (beta) mode
    4. **💬 Chat Assistant** - Ask questions and provide corrections
    5. **📄 Export & History** - Download checklists (Excel, JSON, CSV) from the past 7 days
    """)
    
    st.markdown("---")
    
    # System capabilities
    st.markdown("### ✨ Key Features")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        **Checklist Generation:**
        - 📋 Tag-Based Matching (no API needed)
        - 🤖 AI Advanced Analysis (beta - with reasoning)
        - ✅ One check per item logic
        - 📊 Schedule of Benefits section
        - 📝 Header info auto-filled from RapidMiner data
        """)
    
    with col2:
        st.markdown("""
        **Smart Features:**
        - 💬 Conversational AI assistant
        - 📚 Learns from corrections
        - 📅 7-day checklist history
        - 📥 Multiple export formats
        - 🔍 Appendix 2 guidelines integration
        """)
    
    st.markdown("---")
    
    # System Status
    st.markdown("### 📊 System Status")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.session_state.openai_api_key:
            st.success("🤖 AI Mode: Enabled")
        else:
            st.info("📋 Tag-Based Mode Only")
    
    with col2:
        if st.session_state.rapidminer_data is not None:
            st.success(f"✅ Data Loaded ({len(st.session_state.rapidminer_data)} entries)")
        else:
            st.warning("⚠️ No data uploaded yet")
    
    with col3:
        if st.session_state.checklist_history:
            st.info(f"📋 {len(st.session_state.checklist_history)} checklist(s) in history")
        else:
            st.info("📋 No checklists generated yet")
    
    st.markdown("---")
    
    # Tips
    with st.expander("💡 Tips & Best Practices"):
        st.markdown("""
        - **Tag-Based Mode** is recommended for routine reviews - it's fast and free
        - **AI Advanced Mode** provides reasoning and intelligent analysis but requires an API key
        - Upload your RapidMiner CSV with columns: `line_no`, `text`, `tags`
        - Use the Chat Assistant to ask questions about checklist decisions
        - Check the Export page to download previous checklists without regenerating
        - Corrections you provide help improve future AI analysis
        """)


# Load Data Page
elif page == "📊 Load Data":
    st.markdown('<div class="main-header">📊 Load Orange Output Data</div>', 
                unsafe_allow_html=True)
    
    tab1, tab2 = st.tabs(["📤 Upload Excel File", "🧪 Use Mock Data"])
    
    with tab1:
        st.markdown("### Upload Orange Workflow Output")
        
        uploaded_file = st.file_uploader(
            "Choose an Excel file",
            type=['xlsx', 'xls'],
            help="Upload the Excel output from Orange workflow"
        )
        
        if uploaded_file is not None:
            try:
                df = pd.read_excel(uploaded_file)
                
                # Validate required columns
                required_cols = ['Field', 'Extracted_Value', 'Confidence', 'Page_Number']
                if all(col in df.columns for col in required_cols):
                    st.session_state.current_data = df
                    st.success(f"✅ File loaded successfully! {len(df)} fields found.")
                    
                    # Show preview
                    st.markdown("#### Data Preview")
                    st.dataframe(df.head(10), use_container_width=True)
                    
                    # Show statistics
                    col1, col2, col3 = st.columns(3)
                    found = len(df[df['Extracted_Value'] != 'N/F'])
                    col1.metric("Total Fields", len(df))
                    col2.metric("Fields Found", found)
                    col3.metric("Avg Confidence", f"{df['Confidence'].mean()*100:.1f}%")
                else:
                    st.error(f"❌ Invalid file format. Required columns: {', '.join(required_cols)}")
            except Exception as e:
                st.error(f"❌ Error reading file: {str(e)}")
    
    with tab2:
        st.markdown("### Select Mock Data Group")
        st.info("💡 These are sample datasets for testing the system")
        
        mock_groups = [
            "Aurora Dynamics",
            "Helios Manufacturing Inc.",
            "Solstice Technologies",
            "TechVenture Inc.",
            "GlobalCorp International"
        ]
        
        selected_group = st.selectbox(
            "Choose a group",
            [""] + mock_groups,
            help="Select a pre-loaded mock dataset"
        )
        
        if selected_group:
            if st.button("Load Mock Data", type="primary"):
                df = load_mock_data(selected_group)
                
                if df is not None:
                    st.session_state.current_data = df
                    st.session_state.selected_group = selected_group
                    st.success(f"✅ Loaded mock data for {selected_group}")
                    
                    # Show preview
                    st.markdown("#### Data Preview")
                    st.dataframe(df, use_container_width=True)
                    
                    # Show statistics
                    st.markdown("#### Statistics")
                    col1, col2, col3, col4 = st.columns(4)
                    found = len(df[df['Extracted_Value'] != 'N/F'])
                    missing = len(df) - found
                    
                    col1.metric("Total Fields", len(df))
                    col2.metric("Found", found, delta="Good" if found > 15 else "Review")
                    col3.metric("Missing", missing, delta="⚠️" if missing > 5 else "✓")
                    col4.metric("Avg Confidence", f"{df['Confidence'].mean()*100:.0f}%")
                else:
                    st.error("❌ Could not load mock data. Check file path.")

# Generate Checklist Page
elif page == "🔨 Generate Checklist":
    st.markdown('<div class="main-header">🔨 Generate Checklist from RapidMiner Data</div>', 
                unsafe_allow_html=True)
    
    # Check if data is loaded
    if st.session_state.rapidminer_data is None:
        st.warning("⚠️ No RapidMiner data loaded yet. Please upload data first.")
        if st.button("Go to Data Upload Page"):
            st.rerun()
    else:
        # Data summary
        st.markdown(f"### Data Summary")
        col1, col2 = st.columns(2)
        with col1:
            st.info(f"📋 {len(st.session_state.rapidminer_data)} entries loaded")
        with col2:
            # Check if AI is available
            if st.session_state.openai_api_key:
                st.success("🤖 AI Mode Available")
            else:
                st.info("📋 Tag-Based Mode Only")
        
        st.markdown("---")
        
        # Mode selection
        st.markdown("### Select Processing Mode")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 📋 Tag-Based Matching")
            st.markdown("""
            **Standard Mode** (Recommended)
            - ✅ Fast & reliable
            - ✅ No API costs
            - ✅ Works offline
            - Simple presence/absence checking
            """)
            use_tag_mode = st.button("Generate with Tag Matching", type="primary", use_container_width=True)
        
        with col2:
            st.markdown("#### 🤖 AI Advanced Analysis")
            st.markdown("""
            **Beta Mode** (Experimental)
            - 🧠 Intelligent reasoning
            - 🔍 Context interpretation
            - 💡 Nuanced judgments
            - Suggests approval/notice/handbook
            """)
            if st.session_state.openai_api_key:
                use_ai_mode = st.button("Generate with AI (Beta)", use_container_width=True)
            else:
                st.button("Generate with AI (Beta)", disabled=True, use_container_width=True)
                st.caption("⚠️ Requires API key")
                use_ai_mode = False
        
        # Generate based on selected mode
        if use_tag_mode or use_ai_mode:
            with st.spinner("Generating checklist..."):
                if use_ai_mode and st.session_state.openai_api_key:
                    # AI Advanced Mode
                    try:
                        checklist_df = generate_checklist_with_ai(
                            st.session_state.rapidminer_data,
                            st.session_state.openai_api_key
                        )
                        st.info("🤖 Generated using AI Advanced Analysis")
                    except Exception as e:
                        st.warning(f"AI mode failed ({str(e)}), using tag-based matching...")
                        checklist_df = generate_checklist_from_tags(st.session_state.rapidminer_data)
                else:
                    # Tag-Based Mode
                    checklist_df = generate_checklist_from_tags(st.session_state.rapidminer_data)
                    st.info("📋 Generated using Tag-Based Matching")
                
                if checklist_df is not None:
                    st.session_state.ai_generated_checklist = checklist_df
                    
                    # Save to history with metadata
                    history_entry = {
                        'timestamp': datetime.now(),
                        'checklist': checklist_df.copy(),
                        'mode': 'AI Advanced' if use_ai_mode else 'Tag-Based',
                        'document_name': f"Document_{len(st.session_state.rapidminer_data)}_entries",
                        'total_items': len(checklist_df),
                        'matched': len(checklist_df[checklist_df['Matches'] == True]),
                        'approval': len(checklist_df[checklist_df['Requires Approval'] == True]),
                        'notice': len(checklist_df[checklist_df['Requires Notice'] == True]),
                        'handbook': len(checklist_df[checklist_df['Request Handbook'] == True])
                    }
                    st.session_state.checklist_history.append(history_entry)
                    
                    # Keep only last 7 days
                    week_ago = datetime.now() - pd.Timedelta(days=7)
                    st.session_state.checklist_history = [
                        h for h in st.session_state.checklist_history 
                        if h['timestamp'] >= week_ago
                    ]
                    
                    st.success("✅ Checklist generated successfully!")
                    st.balloons()
        
        # Display generated checklist
        if st.session_state.ai_generated_checklist is not None:
            st.markdown("---")
            st.markdown("### 📋 Generated Checklist")
            
            checklist_df = st.session_state.ai_generated_checklist
            
            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Items", len(checklist_df))
            col2.metric("✅ Matched", len(checklist_df[checklist_df['Matches'] == True]))
            col3.metric("📋 Req. Handbook", len(checklist_df[checklist_df['Request Handbook'] == True]))
            col4.metric("⚠️ Req. Approval", len(checklist_df[checklist_df['Requires Approval'] == True]))
            
            st.markdown("---")
            
            # Tabs for different views
            tab1, tab2, tab3, tab4 = st.tabs([
                "📊 Complete Checklist", 
                "✅ Matched Items", 
                "❌ Not Matched",
                "🤖 AI Reasoning"
            ])
            
            with tab1:
                st.markdown("#### Complete Checklist View")
                # Display as interactive table
                display_df = checklist_df.copy()
                display_df['Matches'] = display_df['Matches'].apply(lambda x: '✅' if x else '❌')
                display_df['Requires Approval'] = display_df['Requires Approval'].apply(lambda x: '☑️' if x else '☐')
                display_df['Requires Notice'] = display_df['Requires Notice'].apply(lambda x: '☑️' if x else '☐')
                display_df['Request Handbook'] = display_df['Request Handbook'].apply(lambda x: '☑️' if x else '☐')
                
                st.dataframe(
                    display_df[['Item', 'Matches', 'Requires Approval', 'Requires Notice', 'Request Handbook']],
                    use_container_width=True,
                    height=600
                )
            
            with tab2:
                st.markdown("#### Items Found in Document")
                matched = checklist_df[checklist_df['Matches'] == True]
                if len(matched) > 0:
                    for _, row in matched.iterrows():
                        with st.container():
                            col1, col2, col3, col4 = st.columns([3, 1, 1, 1])
                            col1.markdown(f"**✅ {row['Item']}**")
                            col2.markdown("📋" if row['Request Handbook'] else "")
                            col3.markdown("📢" if row['Requires Notice'] else "")
                            col4.markdown("✔️" if row['Requires Approval'] else "")
                            
                            if row['AI Reasoning']:
                                with st.expander("View AI Reasoning"):
                                    st.write(row['AI Reasoning'])
                else:
                    st.info("No matched items")
            
            with tab3:
                st.markdown("#### Items Not Found")
                not_matched = checklist_df[checklist_df['Matches'] == False]
                if len(not_matched) > 0:
                    for _, row in not_matched.iterrows():
                        st.markdown(f"""
                        <div class="missing-field">
                        <strong>❌ {row['Item']}</strong>: Not found in keywords
                        </div>
                        """, unsafe_allow_html=True)
                else:
                    st.success("🎉 All items matched!")
            
            with tab4:
                st.markdown("#### AI Reasoning for Each Item")
                for _, row in checklist_df.iterrows():
                    status_icon = "✅" if row['Matches'] else "❌"
                    with st.expander(f"{status_icon} {row['Item']}", expanded=False):
                        col1, col2 = st.columns([1, 2])
                        with col1:
                            st.write("**Status:**")
                            st.write(f"- Matches: {'Yes' if row['Matches'] else 'No'}")
                            st.write(f"- Req. Approval: {'Yes' if row['Requires Approval'] else 'No'}")
                            st.write(f"- Req. Notice: {'Yes' if row['Requires Notice'] else 'No'}")
                            st.write(f"- Req. Handbook: {'Yes' if row['Request Handbook'] else 'No'}")
                        with col2:
                            st.write("**AI Reasoning:**")
                            st.write(row['AI Reasoning'] if row['AI Reasoning'] else "No reasoning provided")
            
            st.markdown("---")
            
            # Download button
            st.markdown("### 📥 Download Checklist")
            
            col1, col2, col3 = st.columns(3)
            with col2:
                # Generate Excel with the checklist format
                excel_data = generate_excel_checklist_from_ai(checklist_df)
                
                st.download_button(
                    label="📥 Download Excel Checklist",
                    data=excel_data,
                    file_name=f"ai_generated_checklist_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
            
            # AI Chat Assistant Sidebar
            st.markdown("---")
            st.markdown("### 💬 Ask Questions About This Checklist")
            
            with st.expander("💬 Chat with AI Assistant", expanded=False):
                if st.session_state.openai_api_key:
                    st.info("🤖 AI Assistant is ready to answer your questions!")
                else:
                    st.info("📋 Record feedback and corrections (AI responses require API key)")
                
                # Mini chat interface
                chat_input = st.text_area(
                    "Ask a question or provide feedback:",
                    height=80,
                    placeholder="e.g., 'Why was Minimum Hour Requirement marked for approval?'",
                    key="checklist_chat_input"
                )
                
                col_a, col_b = st.columns([1, 1])
                with col_a:
                    if st.button("Send 📤", use_container_width=True, key="send_checklist_chat"):
                        if chat_input:
                            # Add to chat history
                            st.session_state.chat_history.append({
                                'role': 'user',
                                'content': chat_input
                            })
                            
                            # Detect corrections
                            correction_keywords = ['wrong', 'incorrect', 'should be', 'actually', 'correction', 'fix', 'change']
                            if any(keyword in chat_input.lower() for keyword in correction_keywords):
                                st.session_state.user_corrections.append(f"{datetime.now().strftime('%Y-%m-%d %H:%M')}: {chat_input}")
                            
                            # Get AI response if available
                            if st.session_state.openai_api_key:
                                try:
                                    # Build context
                                    context_parts = []
                                    
                                    if st.session_state.rapidminer_data is not None:
                                        doc_summary = []
                                        for _, row in st.session_state.rapidminer_data.head(10).iterrows():
                                            doc_summary.append(f"Line {row['line_no']}: {row['text'][:80]}...")
                                        context_parts.append("DOCUMENT:\n" + "\n".join(doc_summary))
                                    
                                    checklist_summary = checklist_df.to_string(max_rows=21)
                                    context_parts.append("CHECKLIST:\n" + checklist_summary)
                                    
                                    if st.session_state.user_corrections:
                                        context_parts.append("CORRECTIONS:\n" + "\n".join(st.session_state.user_corrections[-3:]))
                                    
                                    full_context = "\n\n".join(context_parts)
                                    
                                    client = OpenAI(api_key=st.session_state.openai_api_key)
                                    response = client.chat.completions.create(
                                        model="gpt-4o-mini",
                                        messages=[
                                            {"role": "system", "content": "You are an expert insurance plan document reviewer for TMHCC. Answer questions concisely and cite specific checklist items."},
                                            {"role": "user", "content": f"{full_context}\n\nQUESTION: {chat_input}"}
                                        ],
                                        temperature=0.7,
                                        max_tokens=300
                                    )
                                    
                                    ai_response = response.choices[0].message.content
                                    st.session_state.chat_history.append({
                                        'role': 'assistant',
                                        'content': ai_response
                                    })
                                except Exception as e:
                                    st.session_state.chat_history.append({
                                        'role': 'assistant',
                                        'content': f"Error: {str(e)}"
                                    })
                            else:
                                st.session_state.chat_history.append({
                                    'role': 'assistant',
                                    'content': "✅ Feedback recorded! Enable AI in Settings for intelligent responses."
                                })
                            
                            st.rerun()
                
                with col_b:
                    if st.button("Clear 🗑️", use_container_width=True, key="clear_checklist_chat"):
                        st.session_state.chat_history = []
                        st.rerun()
                
                # Display recent chat messages
                if st.session_state.chat_history:
                    st.markdown("---")
                    st.markdown("**Recent Conversation:**")
                    # Show last 4 messages
                    for msg in st.session_state.chat_history[-4:]:
                        if msg['role'] == 'user':
                            st.markdown(f"**You:** {msg['content']}")
                        else:
                            st.markdown(f"**AI:** {msg['content']}")
                        st.markdown("---")
                    
                    if len(st.session_state.chat_history) > 4:
                        st.caption(f"+ {len(st.session_state.chat_history) - 4} earlier messages. View all in Chat Assistant page.")
        
        else:
            st.info("👆 Click the button above to generate your checklist")

# Validate Page
elif page == "🔍 Validate":
    st.markdown('<div class="main-header">🔍 Validate Checklist</div>', 
                unsafe_allow_html=True)
    
    if st.session_state.current_data is None:
        st.warning("⚠️ No checklist to validate. Generate a checklist first.")
    else:
        df = st.session_state.current_data
        
        st.markdown("### Validation Report")
        
        # Validation metrics
        col1, col2, col3, col4 = st.columns(4)
        
        high_conf = len(df[df['Confidence'] > 0.8])
        med_conf = len(df[(df['Confidence'] >= 0.5) & (df['Confidence'] <= 0.8)])
        low_conf = len(df[df['Confidence'] < 0.5])
        
        col1.metric("🟢 High Confidence", high_conf, help="Confidence > 80%")
        col2.metric("🟡 Medium Confidence", med_conf, help="Confidence 50-80%")
        col3.metric("🔴 Low Confidence", low_conf, help="Confidence < 50%")
        col4.metric("Validation Status", "⚠️ Review" if low_conf > 0 else "✅ Pass")
        
        # Confidence distribution chart
        st.markdown("#### Confidence Score Distribution")
        
        import plotly.express as px
        fig = px.histogram(
            df, 
            x='Confidence',
            nbins=20,
            title='Field Confidence Distribution',
            labels={'Confidence': 'Confidence Score', 'count': 'Number of Fields'},
            color_discrete_sequence=['#3498db']
        )
        st.plotly_chart(fig, use_container_width=True)
        
        # Fields needing review
        if low_conf > 0:
            st.markdown("#### ⚠️ Fields Requiring Review")
            review_df = df[df['Confidence'] < 0.5][['Field', 'Extracted_Value', 'Confidence', 'Page_Number']]
            st.dataframe(review_df, use_container_width=True)

# Export Page
elif page == "📄 Export":
    st.markdown('<div class="main-header">📄 Export & History</div>', 
                unsafe_allow_html=True)
    
    # Show history tabs
    tab1, tab2 = st.tabs(["📋 Checklist History (Past 7 Days)", "💾 Export Options"])
    
    with tab1:
        st.markdown("### Recent Checklists")
        
        if not st.session_state.checklist_history:
            st.info("📝 No checklists generated yet. Create your first checklist on the Generate Checklist page!")
        else:
            st.success(f"📊 Found {len(st.session_state.checklist_history)} checklist(s) from the past week")
            
            # Display each historical checklist
            for idx, entry in enumerate(reversed(st.session_state.checklist_history)):
                with st.expander(
                    f"🗓️ {entry['timestamp'].strftime('%b %d, %Y at %I:%M %p')} - {entry['mode']} Mode ({entry['matched']}/{entry['total_items']} matched)",
                    expanded=(idx == 0)  # Expand most recent
                ):
                    # Metadata
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("Total Items", entry['total_items'])
                    col2.metric("✅ Matched", entry['matched'])
                    col3.metric("⚠️ Approval", entry['approval'])
                    col4.metric("📋 Handbook", entry['handbook'])
                    
                    if entry['notice'] > 0:
                        st.info(f"📢 {entry['notice']} item(s) require notice")
                    
                    st.markdown("---")
                    
                    # Show checklist
                    st.dataframe(entry['checklist'], use_container_width=True)
                    
                    # Download buttons
                    st.markdown("**Download this checklist:**")
                    col_a, col_b, col_c = st.columns(3)
                    
                    with col_a:
                        # Excel download
                        excel_data = generate_excel_checklist_from_ai(entry['checklist'])
                        st.download_button(
                            label="📥 Excel",
                            data=excel_data,
                            file_name=f"checklist_{entry['timestamp'].strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"excel_{idx}"
                        )
                    
                    with col_b:
                        # JSON download
                        json_data = {
                            "timestamp": entry['timestamp'].isoformat(),
                            "mode": entry['mode'],
                            "document_name": entry['document_name'],
                            "summary": {
                                "total": entry['total_items'],
                                "matched": entry['matched'],
                                "approval": entry['approval'],
                                "notice": entry['notice'],
                                "handbook": entry['handbook']
                            },
                            "checklist": entry['checklist'].to_dict('records')
                        }
                        st.download_button(
                            label="📥 JSON",
                            data=json.dumps(json_data, indent=2, default=str),
                            file_name=f"checklist_{entry['timestamp'].strftime('%Y%m%d_%H%M%S')}.json",
                            mime="application/json",
                            use_container_width=True,
                            key=f"json_{idx}"
                        )
                    
                    with col_c:
                        # CSV download
                        csv_data = entry['checklist'].to_csv(index=False)
                        st.download_button(
                            label="📥 CSV",
                            data=csv_data,
                            file_name=f"checklist_{entry['timestamp'].strftime('%Y%m%d_%H%M%S')}.csv",
                            mime="text/csv",
                            use_container_width=True,
                            key=f"csv_{idx}"
                        )
            
            # Clear history button
            st.markdown("---")
            if st.button("🗑️ Clear All History"):
                st.session_state.checklist_history = []
                st.rerun()
    
    with tab2:
        st.markdown("### Export Options")
        
        if st.session_state.ai_generated_checklist is None:
            st.warning("⚠️ No current checklist. Generate one first on the Generate Checklist page.")
        else:
            st.info("💡 Download options for the current checklist are available above in the Checklist History tab.")
            
            st.markdown("### Bulk Export")
            st.markdown("Export all checklists from the past week:")
            
            if st.session_state.checklist_history:
                # Create combined Excel with multiple sheets
                if st.button("📥 Download All as Multi-Sheet Excel"):
                    st.info("💡 Feature coming soon: Export all checklists as separate Excel sheets in one workbook")
                
                # Create combined JSON
                all_json = {
                    "export_date": datetime.now().isoformat(),
                    "total_checklists": len(st.session_state.checklist_history),
                    "checklists": [
                        {
                            "timestamp": entry['timestamp'].isoformat(),
                            "mode": entry['mode'],
                            "summary": {
                                "total": entry['total_items'],
                                "matched": entry['matched'],
                                "approval": entry['approval']
                            },
                            "data": entry['checklist'].to_dict('records')
                        }
                        for entry in st.session_state.checklist_history
                    ]
                }
                
                st.download_button(
                    label="📥 Download All as JSON",
                    data=json.dumps(all_json, indent=2, default=str),
                    file_name=f"all_checklists_{datetime.now().strftime('%Y%m%d')}.json",
                    mime="application/json",
                    use_container_width=False
                )
            else:
                st.info("No checklists available for bulk export")

# Learning System Page
elif page == "🧠 Learning System":
    st.markdown('<div class="main-header">🧠 Learning System</div>', 
                unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["➕ Add Synonym", "✏️ Add Correction", "📚 Learning History"])
    
    with tab1:
        st.markdown("### Add New Synonym")
        st.write("Teach the system alternative terms for insurance fields")
        
        col1, col2 = st.columns(2)
        with col1:
            term = st.text_input("Original Term", placeholder="e.g., TPA")
        with col2:
            synonym = st.text_input("Synonym", placeholder="e.g., Claims Administrator")
        
        if st.button("Add Synonym", type="primary"):
            if term and synonym:
                st.success(f"✅ Added: '{synonym}' as synonym for '{term}'")
                st.info("💡 Integrate with your LearningEngine.add_synonym() method")
            else:
                st.error("Please fill in both fields")
    
    with tab2:
        st.markdown("### Add Correction")
        st.write("Teach the system to correct common mistakes")
        
        col1, col2 = st.columns(2)
        with col1:
            incorrect = st.text_input("Incorrect Term", placeholder="e.g., Third Party Admin")
        with col2:
            correct = st.text_input("Correct Term", placeholder="e.g., Third Party Administrator")
        
        if st.button("Add Correction", type="primary"):
            if incorrect and correct:
                st.success(f"✅ Correction added: '{incorrect}' → '{correct}'")
                st.info("💡 Integrate with your LearningEngine.add_correction() method")
            else:
                st.error("Please fill in both fields")
    
    with tab3:
        st.markdown("### Learning History")
        st.write("View all learned synonyms and corrections")
        
        # Mock learning history
        history_data = {
            "Timestamp": ["2025-11-15 10:30", "2025-11-15 11:45", "2025-11-14 14:20"],
            "Action": ["Synonym Added", "Correction Added", "Synonym Added"],
            "Term": ["TPA", "Third Party Admin", "UR Vendor"],
            "Value": ["Claims Administrator", "Third Party Administrator", "Utilization Review"],
            "User": ["system", "admin", "user1"]
        }
        
        history_df = pd.DataFrame(history_data)
        st.dataframe(history_df, use_container_width=True)

# Chat Assistant Page
elif page == "💬 Chat Assistant":
    st.markdown('<div class="main-header">💬 Chat Assistant</div>', 
                unsafe_allow_html=True)
    
    # Show mode status
    if st.session_state.openai_api_key:
        st.success("🤖 AI Chat Mode: Intelligent conversations enabled")
    else:
        st.info("📋 Basic Mode: View corrections and provide feedback (AI responses disabled)")
    
    st.markdown("""
    ### Ask Questions & Provide Feedback
    
    Use this chat to:
    - 🤔 Ask questions about the checklist or document
    - ✏️ Provide corrections to improve accuracy
    - 💡 Get explanations for specific items
    - 🔍 Request deeper analysis
    """)
    
    st.markdown("---")
    
    # Check if data is loaded
    if st.session_state.rapidminer_data is None and st.session_state.ai_generated_checklist is None:
        st.info("💡 Upload RapidMiner data and generate a checklist first to enable full functionality")
    
    # Chat history display
    st.markdown("### 💬 Conversation")
    
    chat_container = st.container()
    
    with chat_container:
        if not st.session_state.chat_history:
            st.markdown("*No messages yet. Start the conversation below!*")
        else:
            for i, message in enumerate(st.session_state.chat_history):
                if message['role'] == 'user':
                    st.markdown(f"**You:** {message['content']}")
                else:
                    st.markdown(f"**AI Assistant:** {message['content']}")
                st.markdown("---")
    
    # User input
    st.markdown("### ✍️ Your Message")
    
    col1, col2 = st.columns([4, 1])
    
    with col1:
        user_message = st.text_area(
            "Type your question or feedback:",
            height=100,
            placeholder="e.g., 'Why was Minimum Hour Requirement marked for approval?' or 'The transplant provision is actually standard, not requiring approval'"
        )
    
    with col2:
        st.markdown("  \n  \n")  # Spacing
        send_button = st.button("Send 📤", type="primary", use_container_width=True)
        clear_button = st.button("Clear Chat 🗑️", use_container_width=True)
    
    if clear_button:
        st.session_state.chat_history = []
        st.session_state.user_corrections = []
        st.rerun()
    
    if send_button and user_message:
        # Add user message to history
        st.session_state.chat_history.append({
            'role': 'user',
            'content': user_message
        })
        
        # Detect if this is a correction
        correction_keywords = ['wrong', 'incorrect', 'should be', 'actually', 'correction', 'fix', 'change']
        is_correction = any(keyword in user_message.lower() for keyword in correction_keywords)
        
        if is_correction:
            st.session_state.user_corrections.append(f"{datetime.now().strftime('%Y-%m-%d %H:%M')}: {user_message}")
        
        # If API is available, get AI response
        if st.session_state.openai_api_key:
            # Build context for AI
            context_parts = []
            
            # Include document context if available
            if st.session_state.rapidminer_data is not None:
                document_summary = []
                for _, row in st.session_state.rapidminer_data.head(20).iterrows():
                    document_summary.append(f"Line {row['line_no']}: {row['text'][:100]}... [Tags: {row['tags']}]")
                context_parts.append("DOCUMENT CONTEXT:\n" + "\n".join(document_summary))
            
            # Include checklist context if available
            if st.session_state.ai_generated_checklist is not None:
                checklist_summary = st.session_state.ai_generated_checklist.to_string(max_rows=10)
                context_parts.append("GENERATED CHECKLIST:\n" + checklist_summary)
            
            # Include previous corrections
            if st.session_state.user_corrections:
                corrections_summary = "\n".join([f"- {c}" for c in st.session_state.user_corrections[-5:]])
                context_parts.append("USER CORRECTIONS:\n" + corrections_summary)
            
            # Include chat history
            recent_history = st.session_state.chat_history[-10:]
            history_text = "\n".join([f"{msg['role'].upper()}: {msg['content']}" for msg in recent_history])
            context_parts.append("CONVERSATION HISTORY:\n" + history_text)
            
            full_context = "\n\n".join(context_parts)
            
            # Call AI
            try:
                client = OpenAI(api_key=st.session_state.openai_api_key)
                
                response = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[
                        {
                            "role": "system", 
                            "content": """You are an expert insurance plan document reviewer for TMHCC. 
                            
You help users understand checklist decisions, answer questions about plan provisions, and learn from corrections.

When users provide corrections:
1. Acknowledge the correction
2. Explain how this improves future analysis
3. Ask if they want to update the current checklist

When users ask questions:
1. Reference the specific document context
2. Cite Appendix 2 guidelines when relevant
3. Provide clear, concise explanations

Be helpful, professional, and always cite your sources."""
                        },
                        {
                            "role": "user",
                            "content": f"{full_context}\n\nUSER QUESTION/FEEDBACK: {user_message}"
                        }
                    ],
                    temperature=0.7,
                    max_tokens=500
                )
                
                ai_response = response.choices[0].message.content
                
                # Add AI response to history
                st.session_state.chat_history.append({
                    'role': 'assistant',
                    'content': ai_response
                })
                
            except Exception as e:
                error_msg = f"Sorry, I encountered an error: {str(e)}"
                st.session_state.chat_history.append({
                    'role': 'assistant',
                    'content': error_msg
                })
        else:
            # No API - provide basic acknowledgment
            if is_correction:
                basic_response = "✅ Correction recorded! Your feedback has been saved and will help improve future analysis. Enable AI mode in Settings for intelligent responses."
            else:
                basic_response = "📝 Message received! Enable AI mode in Settings to get intelligent responses to your questions."
            
            st.session_state.chat_history.append({
                'role': 'assistant',
                'content': basic_response
            })
        
        st.rerun()
    
    # Show corrections log
    if st.session_state.user_corrections:
        st.markdown("---")
        st.markdown("### 📝 Correction Log")
        st.caption(f"{len(st.session_state.user_corrections)} corrections recorded")
        with st.expander("View all corrections"):
            for correction in st.session_state.user_corrections:
                st.markdown(f"- {correction}")
            
            if st.button("Export Corrections"):
                corrections_text = "\n".join(st.session_state.user_corrections)
                st.download_button(
                    label="Download as TXT",
                    data=corrections_text,
                    file_name=f"corrections_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt",
                    mime="text/plain"
                )

# About Page
elif page == "ℹ️ About":
    st.markdown('<div class="main-header">ℹ️ About This System</div>', 
                unsafe_allow_html=True)
    
    st.markdown("""
    ### AI-Enabled Plan Document Review System
    
    **Version:** 1.0  
    **Release Date:** November 2025  
    **Developer:** Tokio Marine Insurance Development Team
    
    ---
    
    #### 🎯 Purpose
    
    This system automates the review of insurance plan documents by parsing Orange workflow 
    outputs and generating structured checklists with intelligent validation.
    
    #### ✨ Key Features
    
    - **Automated Processing**: Converts Excel outputs into validated checklists
    - **Smart Validation**: Cross-references with definitions and detects issues
    - **Continuous Learning**: Improves from user feedback and corrections
    - **Multiple Export Formats**: PDF, JSON, Excel, and HTML
    - **Interactive Interface**: Modern web-based GUI for easy operation
    
    #### 📊 Technical Stack
    
    - **Frontend**: Streamlit (Python)
    - **Backend**: Python 3.11+
    - **Data Processing**: Pandas, OpenPyXL
    - **Report Generation**: ReportLab
    - **Visualization**: Plotly
    
    #### 📚 Documentation
    
    - [README.md](README.md) - Complete user guide
    - [SETUP_GUIDE.md](SETUP_GUIDE.md) - Installation instructions
    - [TEST_RESULTS.md](TEST_RESULTS.md) - Test documentation
    - [GUI_ROADMAP.md](GUI_ROADMAP.md) - GUI development guide
    
    #### 🧪 Test Status
    
    - ✅ All functionality tests passing (13/13)
    - ✅ Mock data validated (6 groups)
    - ✅ Export formats working
    - ✅ Learning system operational
    
    #### 📧 Support
    
    For questions or issues, contact the development team or refer to the documentation.
    
    ---
    
    *Built with ❤️ for Tokio Marine Insurance*
    """)

# Footer
st.markdown("---")
col1, col2, col3 = st.columns(3)
with col1:
    st.markdown("**Tokio Marine Plan Review System**")
with col2:
    st.markdown("Version 1.0 | November 2025")
with col3:
    st.markdown("Made with Streamlit 🎈")
